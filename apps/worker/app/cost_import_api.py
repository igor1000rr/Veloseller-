"""Импорт себестоимости из CSV/XLSX (массовая загрузка из карточки товара).

POST /cost-prices/import — принимает файл + connection_id + буквы колонок
(артикул / себестоимость), сопоставляет товары по SKU в пределах склада и
проставляет products.cost_price. Возвращает {matched, totalRows, unmatched}.

Вызывается из /api/cost-prices/import (Next.js → Worker).
Аутентификация — X-Worker-Secret (вешается через include_router в main.py).

Колонки указываются буквами Excel (A, B, ... D, F), как в спецификации
Александра — поэтому парсим файл позиционно, а не по заголовкам (в отличие
от radar/price_parser.py, который ключует по первой строке).
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any, Optional

from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from app.db import get_supabase

logger = logging.getLogger("veloseller.worker.cost_import")

router = APIRouter(prefix="/cost-prices", tags=["cost-prices"])

# Защита от гигантских файлов (строк данных). 100k SKU — заведомо с запасом.
_MAX_ROWS = 100_000
# Размер пачки для bulk-RPC, чтобы jsonb-payload не разрастался.
_RPC_BATCH = 5_000


def _excel_col_to_index(letter: str) -> int:
    """Буква колонки Excel → 0-based индекс. 'A'→0, 'D'→3, 'F'→5, 'AA'→26."""
    s = (letter or "").strip().upper()
    if not re.fullmatch(r"[A-Z]+", s):
        raise ValueError(f"Некорректная буква колонки: {letter!r} (ожидается A, B, C…)")
    idx = 0
    for ch in s:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _parse_grid(file_bytes: bytes, file_name: str) -> list[list[Any]]:
    """Парсит XLSX/XLS/CSV в позиционную сетку (список строк-списков)."""
    ext = file_name.lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover
            raise RuntimeError("openpyxl не установлен — нельзя парсить XLSX") from e
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []
        return [list(row) for row in ws.iter_rows(values_only=True)]

    if ext == "csv":
        import csv
        text = file_bytes.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [list(r) for r in csv.reader(io.StringIO(text), dialect=dialect)]

    raise RuntimeError(f"Формат .{ext} не поддерживается. Используйте XLSX или CSV.")


def _parse_cost(raw: Any) -> Optional[float]:
    """Нормализует ячейку себестоимости в float.

    Поддерживает '1 234,56', '1234.56', '1 234 ₽' и т.п. Отрицательные и
    нечисловые → None (строка просто не учитывается).
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        v = float(raw)
        return v if v >= 0 else None
    s = str(raw).strip()
    if not s:
        return None
    # Убираем неразрывные пробелы/пробелы/валюту, запятую → точка.
    s = s.replace("\xa0", "").replace(" ", "").replace("₽", "").replace("руб", "")
    s = s.replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        v = float(m.group(0))
    except ValueError:
        return None
    return v if v >= 0 else None


def build_cost_map(grid: list[list[Any]], art_idx: int, cost_idx: int) -> dict[str, float]:
    """Собирает {артикул: себестоимость} из сетки.

    Заголовок не детектируем намеренно: строка с заголовком просто не
    сопоставится ни с одним SKU. Последнее значение по артикулу побеждает.
    """
    out: dict[str, float] = {}
    for row in grid:
        if art_idx >= len(row):
            continue
        art_raw = row[art_idx]
        if art_raw is None:
            continue
        art = str(art_raw).strip()
        if not art:
            continue
        cost = _parse_cost(row[cost_idx]) if cost_idx < len(row) else None
        if cost is None:
            continue
        out[art] = cost
    return out


@router.post("/import")
async def import_cost_prices(
    seller_id: str = Form(...),
    connection_id: str = Form(...),
    article_col: str = Form(...),
    cost_col: str = Form(...),
    file: UploadFile = File(...),
) -> dict:
    """Сопоставить товары по артикулу и проставить себестоимость для склада."""
    sb = get_supabase()

    try:
        art_idx = _excel_col_to_index(article_col)
        cost_idx = _excel_col_to_index(cost_col)
    except ValueError as e:
        raise HTTPException(400, str(e))

    file_bytes = await file.read()
    file_name = file.filename or "costs.xlsx"

    logger.info("cost_import start", extra={
        "seller_id": seller_id, "connection_id": connection_id,
        "file_name": file_name, "size": len(file_bytes),
        "article_col": article_col, "cost_col": cost_col,
    })

    try:
        grid = _parse_grid(file_bytes, file_name)
    except Exception as e:
        logger.exception("cost_import parse failed", extra={"seller_id": seller_id})
        raise HTTPException(400, f"Не удалось прочитать файл: {e}")

    if len(grid) > _MAX_ROWS:
        raise HTTPException(413, f"Слишком много строк: {len(grid)} > {_MAX_ROWS}")

    file_costs = build_cost_map(grid, art_idx, cost_idx)
    total_rows = len(file_costs)
    if total_rows == 0:
        raise HTTPException(
            422,
            "В файле не найдено пар «артикул → себестоимость». "
            "Проверьте буквы колонок и формат файла.",
        )

    # Bulk-UPDATE через RPC пачками — один UPDATE из jsonb вместо построчных.
    items = list(file_costs.items())
    matched = 0
    for i in range(0, len(items), _RPC_BATCH):
        chunk = items[i:i + _RPC_BATCH]
        payload = [{"sku": sku, "cost": cost} for sku, cost in chunk]
        try:
            res = sb.rpc("bulk_update_cost_prices", {
                "p_seller_id": seller_id,
                "p_connection_id": connection_id,
                "p_costs": payload,
            }).execute()
            matched += int(getattr(res, "data", 0) or 0)
        except Exception as e:
            logger.exception("bulk_update_cost_prices RPC failed",
                             extra={"seller_id": seller_id})
            raise HTTPException(500, f"Ошибка обновления себестоимости: {e}")

    logger.info("cost_import done", extra={
        "seller_id": seller_id, "connection_id": connection_id,
        "file_rows": total_rows, "matched": matched,
    })
    return {
        "totalRows": total_rows,
        "matched": matched,
        "unmatched": max(0, total_rows - matched),
    }
