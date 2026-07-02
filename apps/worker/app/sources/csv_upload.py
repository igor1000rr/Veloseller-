"""CSV upload source.

Формат: колонки sku, product_name (опционально), stock_quantity, price.

БАГ 16 fix: utf-8-sig снимает BOM в начале файла. Excel экспортирует CSV с BOM,
из-за чего первая колонка заголовка превращалась в '\\ufeffsku' и required check проваливался.
БАГ 56 fix: лимит на 50K строк (защита от OOM).
БАГ 59 fix: per-row error handling — одна битая строка не валит весь impport.
БАГ 60 fix: дедупликация SKU внутри одного файла (последняя запись побеждает).
"""
from __future__ import annotations
import csv
import io
import logging
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Iterable
from app.schemas import SnapshotInput

logger = logging.getLogger("veloseller.csv")

# 50K SKU — больше чем у любого реального селлера. Защита от OOM.
MAX_ROWS = 50_000


def parse_csv(content: bytes | str) -> list[SnapshotInput]:
    # БАГ 16: utf-8-sig чтобы убрать BOM от Excel-экспортированных CSV.
    # Если в начале файла \ufeff — питон автоматически срежет.
    if isinstance(content, bytes):
        text = content.decode("utf-8-sig")
    else:
        text = content.lstrip("\ufeff")  # на случай если уже decoded
    reader = csv.DictReader(io.StringIO(text))
    required = {"sku", "stock_quantity", "price"}
    if not required.issubset({h.strip().lower() for h in (reader.fieldnames or [])}):
        raise ValueError(f"CSV должен содержать колонки {required}")

    now = datetime.now(timezone.utc)
    seen_skus: dict[str, SnapshotInput] = {}
    errors: list[str] = []
    row_count = 0
    for row_idx, row in enumerate(reader, start=2):
        row_count += 1
        if row_count > MAX_ROWS:
            raise ValueError(f"CSV содержит более {MAX_ROWS} строк — слишком большой файл")

        try:
            norm = {k.strip().lower(): v for k, v in row.items() if k}
            sku = (norm.get("sku") or "").strip()
            if not sku:
                errors.append(f"строка {row_idx}: пустой SKU")
                continue

            stock_raw = (norm.get("stock_quantity") or "").strip()
            if not stock_raw:
                errors.append(f"строка {row_idx} ({sku}): пустой stock_quantity")
                continue

            try:
                stock = int(float(stock_raw))
            except (ValueError, TypeError):
                errors.append(f"строка {row_idx} ({sku}): невалидный stock_quantity '{stock_raw}'")
                continue
            if stock < 0:
                errors.append(f"строка {row_idx} ({sku}): отрицательный stock {stock}")
                continue

            price_raw = (norm.get("price") or "").strip().replace(",", ".")
            try:
                price = Decimal(price_raw) if price_raw else Decimal("0")
            except (InvalidOperation, ValueError):
                errors.append(f"строка {row_idx} ({sku}): невалидная цена '{price_raw}'")
                continue
            if price < 0:
                errors.append(f"строка {row_idx} ({sku}): отрицательная цена {price}")
                continue

            seen_skus[sku] = SnapshotInput(
                sku=sku,
                product_name=(norm.get("product_name") or "").strip() or None,
                stock_quantity=stock,
                price=price,
                snapshot_time=now,
            )
        except Exception as e:
            errors.append(f"строка {row_idx}: {e}")
            continue

    if errors:
        logger.warning("CSV parsing had errors", extra={
            "errors_count": len(errors),
            "first_errors": errors[:10],
            "successful_rows": len(seen_skus),
        })

    if not seen_skus:
        raise ValueError(
            f"Из {row_count} строк ни одна не прошла валидацию. "
            f"Первые ошибки: {'; '.join(errors[:5])}"
        )

    return list(seen_skus.values())


def parse_xlsx(content: bytes) -> list[SnapshotInput]:
    """Разбор Excel .xlsx (первый лист) через конвертацию в CSV → parse_csv.

    Переиспользуем всю валидацию/дедуп/обработку BOM из parse_csv, не дублируя
    логику. openpyxl уже в зависимостях воркера. Читаем в read_only + data_only
    (кэшированные значения формул), первая непустая строка = заголовки.
    """
    from openpyxl import load_workbook  # локальный импорт — не тянем при обычном CSV

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Не удалось открыть .xlsx: {e}")
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("В книге нет активного листа")
        out = io.StringIO()
        writer = csv.writer(out)
        wrote = 0
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            cells = ["" if c is None else str(c) for c in row]
            if not any(c.strip() for c in cells):
                continue  # пропускаем полностью пустые строки
            writer.writerow(cells)
            wrote += 1
            if wrote > MAX_ROWS + 1:  # +1 на заголовок; parse_csv добьёт точный лимит
                break
    finally:
        wb.close()

    text = out.getvalue()
    if not text.strip():
        raise ValueError("Файл .xlsx пустой — нет строк с данными")
    return parse_csv(text)
