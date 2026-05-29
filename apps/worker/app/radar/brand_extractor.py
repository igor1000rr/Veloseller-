"""Извлечение брендов из прайса через DeepSeek (OpenAI-совместимый API).

Workflow:
  1. Парсим XLSX/CSV → list[dict[colname, value]]
  2. Берём первые N строк (max 200) — достаточно чтобы AI понял структуру и
     увидел разнообразие брендов
  3. Шлём AI-промт: "вот строки прайса, верни JSON-список брендов которые
     ты увидел"
  4. Парсим ответ, нормализуем, проверяем лимит тарифа
  5. Сохраняем в radar_brands со status=approved
  6. Возвращаем (brands_list, ai_metrics) для записи в radar_price_uploads

ENV переменные (priority order):
  DEEPSEEK_API_KEY   — основной (рекомендуется), прямой API api.deepseek.com
  OPENROUTER_API_KEY — fallback, через OpenRouter (если ходим из-под VPN)

DEEPSEEK_MODEL — необязательно, default 'deepseek-chat' (DeepSeek V3, дёшево)
                 альтернатива 'deepseek-reasoner' (R1, для сложных задач)

История: 29.05.2026 переход с OpenRouter на DeepSeek прямой — OpenRouter
перестал принимать платежи из РФ. DeepSeek работает напрямую (китайский
сервис), не блокируется, формат OpenAI-совместимый.
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
from dataclasses import dataclass, field
from typing import Any, Optional

import requests

logger = logging.getLogger("veloseller.radar.brand_extractor")

# DeepSeek основной (рекомендуется в РФ), OpenRouter fallback
_DEEPSEEK_URL = "https://api.deepseek.com/v1/chat/completions"
_OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

_REQUEST_TIMEOUT_SEC = 120.0
_MAX_ROWS_TO_AI = 200  # больше не отправляем — это уже >30K input tokens

# Дефолтная модель DeepSeek — V3 chat ($0.27/M in, $1.10/M out)
# В 3-4 раза дешевле Claude Haiku, качество на task «бренды из прайса» сравнимое
_DEFAULT_DEEPSEEK_MODEL = "deepseek-chat"
_DEFAULT_OPENROUTER_MODEL = "anthropic/claude-haiku-4.5"

# Системный промт для AI: оно должно вернуть строго JSON
_SYSTEM_PROMPT = """Ты — помощник для селлеров маркетплейсов (OZON, Wildberries).
Тебе дают строки прайса селлера. Твоя задача — найти ВСЕ уникальные бренды/марки производителей в этих строках.

Бренды бывают разные:
- Иностранные: Dyson, Samsung, Bosch, Apple, Xiaomi, Philips, LG, Sony, Adidas, Nike
- Российские: Технодом, ИЛЬ ДЕ БОТЭ, Шерсть Бабушки, Простоквашино
- Узкоспециализированные: SKF (подшипники), STANLEY (инструмент)

Игнорируй:
- Общие категории: "молоко", "телевизор", "ноутбук"
- Артикулы: "ABC-123", "V11-SV15"
- Размеры/характеристики: "Black", "32GB", "XL"
- Названия моделей: "iPhone 15" → бренд "Apple" (НЕ "iPhone")

ВАЖНО:
- Если в строке "Сухой пылесос Dyson V12 Detect Slim Absolute" — бренд "Dyson"
- Если в строке "Bosch GBH 2-26 DRE Professional" — бренд "Bosch"
- Если бренд явно не виден — пропусти строку, не выдумывай

Верни ТОЛЬКО валидный JSON без пояснений в формате:
{
  "brands": ["Dyson", "Bosch", "Samsung", ...]
}

Если ни одного бренда не нашёл — верни {"brands": []}.
"""


@dataclass
class ExtractedBrand:
    name: str
    sku_count: int = 0  # сколько строк прайса содержат этот бренд


@dataclass
class BrandExtractionResult:
    brands: list[ExtractedBrand] = field(default_factory=list)
    ai_model: str = ""
    ai_input_tokens: int = 0
    ai_output_tokens: int = 0
    ai_cost_usd: float = 0.0
    ai_raw_response: dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None


def parse_price_file(file_bytes: bytes, file_name: str) -> list[dict[str, Any]]:
    """Парсит XLSX/XLS/CSV в список словарей.

    Возвращает список dict — по одному на строку. Колонки берутся из первой
    строки (header). Пустые ячейки → None.
    """
    ext = file_name.lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError("openpyxl не установлен — нельзя парсить XLSX") from e
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else f"col{i}"
                   for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            d = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    d[headers[i]] = val
            result.append(d)
        return result

    if ext == "csv":
        import csv
        text = file_bytes.decode("utf-8-sig", errors="replace")
        # Авто-определяем разделитель
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        return [dict(r) for r in reader]

    raise RuntimeError(f"Формат .{ext} не поддерживается. Используй XLSX или CSV.")


def _sample_rows_for_ai(rows: list[dict], max_rows: int = _MAX_ROWS_TO_AI) -> list[dict]:
    """Берёт репрезентативную выборку строк для AI.

    Стратегия: первые 50, последние 50, и случайные из середины. Это даёт
    разнообразие если прайс отсортирован (часто бывает: сначала бренд A,
    потом B, и т.д.).
    """
    if len(rows) <= max_rows:
        return rows
    head = rows[:50]
    tail = rows[-50:]
    middle_start = 50
    middle_end = len(rows) - 50
    if middle_end <= middle_start:
        return head + tail
    # равномерно из середины
    needed = max_rows - 100
    step = max(1, (middle_end - middle_start) // needed)
    middle = rows[middle_start:middle_end:step][:needed]
    return head + middle + tail


def _format_rows_as_text(rows: list[dict]) -> str:
    """Превращает строки прайса в компактный текст для AI."""
    lines = []
    for i, row in enumerate(rows, 1):
        # Берём только string-значения, склеиваем через пробел
        parts = []
        for k, v in row.items():
            if v is None or v == "":
                continue
            v_str = str(v).strip()
            if v_str and not v_str.isdigit():  # пропускаем чистые числа
                parts.append(v_str)
        if parts:
            line = " | ".join(parts)
            lines.append(f"{i}. {line[:300]}")  # cap длины строки
    return "\n".join(lines)


def _calculate_cost(model: str, input_tokens: int, output_tokens: int) -> float:
    """Стоимость в USD по тарифу провайдера.

    Цены за 1M токенов (актуально на май 2026):

    DeepSeek:
      deepseek-chat:     $0.27/$1.10  (V3 — дешёвая universal модель)
      deepseek-reasoner: $0.55/$2.19  (R1 — reasoning, для сложных задач)

    OpenRouter (если используется как fallback):
      anthropic/claude-haiku-4.5:  $1.00/$5.00
      anthropic/claude-sonnet-4.6: $3.00/$15.00
      anthropic/claude-opus-4.7:   $5.00/$25.00
    """
    rates = {
        # DeepSeek прямой
        "deepseek-chat":      (0.27, 1.10),
        "deepseek-reasoner":  (0.55, 2.19),
        # OpenRouter — Claude
        "anthropic/claude-haiku-4.5":  (1.0, 5.0),
        "anthropic/claude-sonnet-4.6": (3.0, 15.0),
        "anthropic/claude-opus-4.7":   (5.0, 25.0),
    }
    in_rate, out_rate = rates.get(model, (0.27, 1.10))  # default deepseek-chat
    return (input_tokens * in_rate + output_tokens * out_rate) / 1_000_000


def _resolve_provider() -> tuple[str, str, str, dict[str, str]]:
    """Определяет какой провайдер использовать на основе env переменных.

    Returns:
        (api_url, api_key, default_model, extra_headers)

    Приоритет:
      1. DEEPSEEK_API_KEY — основной (рекомендуется в РФ)
      2. OPENROUTER_API_KEY — fallback (если работает VPN или прокси)

    Если ни один не задан — возвращает пустой api_key, вызывающий код вернёт
    BrandExtractionResult с error.
    """
    deepseek_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    openrouter_key = os.getenv("OPENROUTER_API_KEY", "").strip()

    if deepseek_key:
        return (
            _DEEPSEEK_URL,
            deepseek_key,
            os.getenv("DEEPSEEK_MODEL", _DEFAULT_DEEPSEEK_MODEL),
            {},  # DeepSeek не требует никаких доп. заголовков
        )

    if openrouter_key:
        return (
            _OPENROUTER_URL,
            openrouter_key,
            os.getenv("OPENROUTER_MODEL", _DEFAULT_OPENROUTER_MODEL),
            {
                "HTTP-Referer": "https://veloseller.ru",
                "X-Title": "Veloseller Radar",
            },
        )

    return ("", "", "", {})


def extract_brands_from_price(
    file_bytes: bytes,
    file_name: str,
    *,
    max_brands: int = 100,
) -> BrandExtractionResult:
    """Главная точка входа: парсит прайс → AI → список брендов.

    Не сохраняет в БД — это делает вызывающий код в API-роуте.
    Используется и в /api/radar/upload, и в worker если будет отложенная
    обработка.
    """
    api_url, api_key, model, extra_headers = _resolve_provider()
    if not api_key:
        return BrandExtractionResult(
            error="Ни DEEPSEEK_API_KEY, ни OPENROUTER_API_KEY не настроены в env"
        )

    # 1. Парсим файл
    try:
        rows = parse_price_file(file_bytes, file_name)
    except Exception as e:
        return BrandExtractionResult(error=f"Ошибка парсинга файла: {e}")

    if not rows:
        return BrandExtractionResult(error="Прайс пустой или нет ни одной строки данных")

    # 2. Выборка строк
    sampled = _sample_rows_for_ai(rows)
    rows_text = _format_rows_as_text(sampled)
    if not rows_text:
        return BrandExtractionResult(
            error="В прайсе не нашлось текстовых данных — нечего анализировать"
        )

    # 3. Запрос к AI (OpenAI-совместимый формат, работает и для DeepSeek и для OpenRouter)
    try:
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            **extra_headers,
        }
        resp = requests.post(
            api_url,
            headers=headers,
            json={
                "model": model,
                "messages": [
                    {"role": "system", "content": _SYSTEM_PROMPT},
                    {"role": "user", "content": f"Строки прайса:\n\n{rows_text}\n\nВерни JSON со списком брендов."},
                ],
                "temperature": 0.0,
                "max_tokens": 4000,
                # DeepSeek поддерживает response_format для гарантии JSON,
                # но не все версии. Не используем, парсим вручную с regex.
            },
            timeout=_REQUEST_TIMEOUT_SEC,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        return BrandExtractionResult(error=f"AI API ошибка: {e}")

    # 4. Парсим ответ
    try:
        content = data["choices"][0]["message"]["content"]
    except (KeyError, IndexError) as e:
        return BrandExtractionResult(
            error=f"Неожиданный формат ответа AI: {e}",
            ai_raw_response=data,
        )

    # Извлекаем JSON (модель может обернуть в ```json ... ```)
    json_match = re.search(r"\{[^{}]*\"brands\"[^{}]*\}", content, re.DOTALL)
    brands_json_str = json_match.group(0) if json_match else content
    try:
        parsed = json.loads(brands_json_str)
        ai_brand_names = parsed.get("brands", []) or []
    except json.JSONDecodeError:
        return BrandExtractionResult(
            error=f"Не удалось разобрать JSON от AI: {content[:300]}",
            ai_raw_response=data,
        )

    if not isinstance(ai_brand_names, list):
        return BrandExtractionResult(
            error=f"AI вернул не-список брендов: {type(ai_brand_names).__name__}",
            ai_raw_response=data,
        )

    # 5. Нормализация и подсчёт SKU
    seen_normalized: set[str] = set()
    extracted: list[ExtractedBrand] = []
    for raw_name in ai_brand_names:
        name = str(raw_name).strip()
        if not name or len(name) < 2 or len(name) > 100:
            continue
        normalized = name.lower()
        if normalized in seen_normalized:
            continue
        seen_normalized.add(normalized)
        # Считаем сколько строк прайса упоминают этот бренд
        sku_count = sum(
            1 for row in rows
            if any(normalized in str(v).lower() for v in row.values() if v is not None)
        )
        extracted.append(ExtractedBrand(name=name, sku_count=sku_count))

    # Сортировка по sku_count убыванию (популярные бренды первыми)
    extracted.sort(key=lambda b: b.sku_count, reverse=True)
    if max_brands > 0:
        extracted = extracted[:max_brands]

    # 6. Метрики стоимости
    usage = data.get("usage", {}) or {}
    input_tokens = int(usage.get("prompt_tokens", 0) or 0)
    output_tokens = int(usage.get("completion_tokens", 0) or 0)
    cost_usd = _calculate_cost(model, input_tokens, output_tokens)

    return BrandExtractionResult(
        brands=extracted,
        ai_model=model,
        ai_input_tokens=input_tokens,
        ai_output_tokens=output_tokens,
        ai_cost_usd=cost_usd,
        ai_raw_response=data,
    )
