"""Парсер прайса XLSX/XLS/CSV → list[dict].

Выделен из brand_extractor.py 29.05.2026. Было важно разделить парсинг файла
(это инфраструктурный код) и извлечение брендов (бизнес-логика). Их
жизненные циклы разные и изменяются по разным причинам.
"""
from __future__ import annotations

import io
from typing import Any


def parse_price_file(file_bytes: bytes, file_name: str) -> list[dict[str, Any]]:
    """Парсит XLSX/XLS/CSV в список словарей.

    Возвращает список dict — по одному на строку. Колонки берутся из первой
    строки (header). Пустые ячейки → None.
    """
    ext = file_name.lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError("openpyxl не установлен — нельзя парсить XLSX") from e
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else f"col{i}"
                   for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            d = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    d[headers[i]] = val
            result.append(d)
        return result

    if ext == "csv":
        import csv
        text = file_bytes.decode("utf-8-sig", errors="replace")
        # Авто-определяем разделитель
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        return [dict(r) for r in reader]

    raise RuntimeError(f"Формат .{ext} не поддерживается. Используй XLSX или CSV.")
