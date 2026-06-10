"""Еженедельный Excel-отчёт по складам — email и Telegram.

Каждый понедельник в 09:00 UTC (12:00 МСК) генерируем для каждого активного
селлера (plan ≠ trial ИЛИ имеющего склады) Excel с тремя листами:
  “Сводка” — store_metrics 30 дней
  “Топ потерь” — SKU с большим lost_revenue
  “Неликвид” — SKU с coverage_days > 180
Отправляем через Resend (attachment) и Telegram (sendDocument), респектим notify_email / notify_telegram.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta, timezone
from typing import Optional

from app.db import fetch_all, get_supabase

logger = logging.getLogger("veloseller.weekly_report")


def _format_money(value, currency: str = "RUB") -> str:
    if value is None:
        return "—"
    try:
        num = float(value)
    except (TypeError, ValueError):
        return "—"
    sign = "₽" if currency == "RUB" else currency
    return f"{num:,.0f} {sign}".replace(",", " ")


def _bold(cell, color="0F172A") -> None:
    from openpyxl.styles import Font
    cell.font = Font(bold=True, color=color)


def _column_widths(ws, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _build_summary_sheet(wb, store_metrics: list[dict], currency: str) -> None:
    """Лист 1: сводная метрика за последние 14 периодов store_metrics."""
    ws = wb.create_sheet("Сводка")
    headers = [
        "Дата", "Всего SKU", "Нет в наличии", "Низкий остаток",
        "Неликвид (SKU)", "Health Score", "Остатки (деньги)",
        "Заморожено в неликвиде", "Потерянная выручка",
        "Конц. остатков 50%", "Конц. спроса 50%",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _bold(ws.cell(row=1, column=col_idx))

    for m in store_metrics[:14]:
        period_end = (m.get("period_end") or "")[:10]
        ws.append([
            period_end,
            m.get("total_sku_count") or 0,
            m.get("oos_sku_count") or 0,
            m.get("low_stock_sku_count") or 0,
            m.get("dead_inventory_sku_count") or 0,
            round(float(m.get("warehouse_health_score") or 0), 1),
            _format_money(m.get("total_inventory_value"), currency),
            _format_money(m.get("store_frozen_inventory_value"), currency),
            _format_money(m.get("lost_revenue"), currency),
            m.get("inventory_concentration_50") or 0,
            m.get("demand_concentration_50") or 0,
        ])

    _column_widths(ws, {
        "A": 12, "B": 12, "C": 14, "D": 14, "E": 14, "F": 13,
        "G": 22, "H": 22, "I": 22, "J": 18, "K": 18,
    })
    ws.freeze_panes = "A2"


def _build_top_losses_sheet(wb, top_losses: list[dict], currency: str) -> None:
    """Лист 2: SKU с самыми большими lost_revenue за последний период."""
    ws = wb.create_sheet("Топ потерь")
    headers = [
        "SKU", "Название", "Скорость продаж (шт/день)",
        "Дней остатка", "Потерянная выручка", "Сегмент",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _bold(ws.cell(row=1, column=col_idx))

    for row in top_losses[:50]:
        products = row.get("products") or {}
        if isinstance(products, list):
            products = products[0] if products else {}
        ws.append([
            products.get("sku") or "—",
            products.get("product_name") or "—",
            round(float(row.get("adjusted_velocity") or 0), 2),
            int(row.get("coverage_days") or 0),
            _format_money(row.get("lost_revenue"), currency),
            row.get("demand_pattern") or "—",
        ])

    _column_widths(ws, {"A": 22, "B": 40, "C": 22, "D": 16, "E": 22, "F": 18})
    ws.freeze_panes = "A2"


def _build_dead_inventory_sheet(wb, dead: list[dict], currency: str) -> None:
    """Лист 3: неликвид (coverage > 180)."""
    ws = wb.create_sheet("Неликвид")
    headers = [
        "SKU", "Название", "Дней остатка",
        "Скорость продаж (шт/день)", "Заморожено",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        _bold(ws.cell(row=1, column=col_idx))

    for row in dead[:200]:
        products = row.get("products") or {}
        if isinstance(products, list):
            products = products[0] if products else {}
        ws.append([
            products.get("sku") or "—",
            products.get("product_name") or "—",
            (int(row["coverage_days"]) if row.get("coverage_days") is not None else "∞"),
            round(float(row.get("adjusted_velocity") or 0), 2),
            _format_money(
                float(row.get("current_stock") or 0) * float(row.get("current_price") or 0),
                currency,
            ),
        ])

    _column_widths(ws, {"A": 22, "B": 40, "C": 16, "D": 22, "E": 22})
    ws.freeze_panes = "A2"


def _generate_excel_for_seller(sb, seller_id: str, currency: str = "RUB") -> Optional[bytes]:
    """Генерирует Excel отчёт в памяти. Возвращает None если данных нет или ошибка."""
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.error("openpyxl not installed — weekly report skipped")
        return None

    try:
        store_metrics_res = (
            sb.table("store_metrics")
            .select("*")
            .eq("seller_id", seller_id)
            .order("period_end", desc=True)
            .limit(14)
            .execute()
        )
        store_metrics = store_metrics_res.data or []

        if not store_metrics:
            logger.info("weekly_report skip seller %s — no store_metrics", seller_id)
            return None

        # Топ-50 потерь из tvelo_metrics за последний период
        top_losses_res = (
            sb.table("tvelo_metrics")
            .select("adjusted_velocity,coverage_days,lost_revenue,demand_pattern,products!inner(sku,product_name,seller_id)")
            .eq("products.seller_id", seller_id)
            .order("lost_revenue", desc=True)
            .limit(50)
            .execute()
        )
        top_losses = top_losses_res.data or []

        # Неликвид: сегмент dead_inventory_risk (coverage > 180 + мёртвый по скорости).
        # Раньше селектили несуществующую колонку frozen_inventory_value — запрос падал.
        dead_res = (
            sb.table("tvelo_metrics")
            .select("coverage_days,adjusted_velocity,current_stock,current_price,inventory_segment,products!inner(sku,product_name,seller_id)")
            .eq("products.seller_id", seller_id)
            .eq("inventory_segment", "dead_inventory_risk")
            .order("current_stock", desc=True)
            .limit(200)
            .execute()
        )
        dead = dead_res.data or []
    except Exception:
        logger.exception("weekly_report data fetch failed for %s", seller_id)
        return None

    wb = Workbook()
    # Убираем дефолтный Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_summary_sheet(wb, store_metrics, currency)
    _build_top_losses_sheet(wb, top_losses, currency)
    _build_dead_inventory_sheet(wb, dead, currency)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_weekly_reports() -> None:
    """Схедулер-джоб. Для каждого селлера с notify_email/telegram подпиской и хотя бы 1 складом — шлём отчёт."""
    try:
        sb = get_supabase()
        sellers = fetch_all(
            sb.table("sellers").select("id,email,display_name,currency,plan,telegram_chat_id,notify_email,notify_telegram")
        )
        sent_email = 0
        sent_telegram = 0
        skipped = 0
        for s in sellers:
            seller_id = s["id"]
            try:
                # Скип если нет складов вообще (пустые аккаунты)
                conn_count_res = (
                    sb.table("data_connections").select("id", count="exact", head=True)
                    .eq("seller_id", seller_id).limit(1).execute()
                )
                if not getattr(conn_count_res, "count", 0):
                    skipped += 1
                    continue

                # Скип если выключены оба канала
                if not s.get("notify_email") and not s.get("notify_telegram"):
                    skipped += 1
                    continue

                currency = s.get("currency") or "RUB"
                xlsx_bytes = _generate_excel_for_seller(sb, seller_id, currency)
                if not xlsx_bytes:
                    skipped += 1
                    continue

                today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                filename = f"veloseller-report-{today}.xlsx"

                # Email
                if s.get("notify_email") and s.get("email"):
                    try:
                        from app.notifications import send_weekly_report_email
                        if send_weekly_report_email(
                            to_email=s["email"],
                            seller_name=s.get("display_name"),
                            xlsx_bytes=xlsx_bytes,
                            filename=filename,
                        ):
                            sent_email += 1
                    except Exception:
                        logger.exception("weekly_report email failed for %s", seller_id)

                # Telegram
                if s.get("notify_telegram") and s.get("telegram_chat_id"):
                    try:
                        from app.telegram import send_document
                        if send_document(
                            chat_id=s["telegram_chat_id"],
                            file_bytes=xlsx_bytes,
                            filename=filename,
                            caption="📊 <b>Veloseller — еженедельный отчёт</b>\n\nСводные метрики, топ-50 потерь и неликвид.",
                        ):
                            sent_telegram += 1
                    except Exception:
                        logger.exception("weekly_report telegram failed for %s", seller_id)

            except Exception:
                logger.exception("weekly_report failed for seller %s", seller_id)
                continue

        logger.info("weekly_report job done", extra={
            "total_sellers": len(sellers),
            "sent_email": sent_email,
            "sent_telegram": sent_telegram,
            "skipped": skipped,
        })
    except Exception:
        logger.exception("weekly_report job crashed")
