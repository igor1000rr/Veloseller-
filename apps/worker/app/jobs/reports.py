"""Диспетчер Excel-отчётов: daily + weekly.

Архитектура (рефакторинг 01.06.2026 — правки Александра из Veloseller_Отчёт.txt):

Стратегия Александра по разделению отчётов:
- ЕЖЕНЕДЕЛЬНЫЕ — Excel операционный, пользователь сам регулирует подписки
- МЕСЯЧНЫЕ     — PDF/Word управленческий, шлётся автоматически 1-го числа.
                  Реализован в apps/worker/app/jobs/monthly_report.py

Состав листов еженедельного Excel (в этом порядке):
1. Сводка по складу         (kind=weekly_report)    — HEAD-страница с числами
2. Потерянные продажи       (kind=underestimated_sku) — TVelo×OOS×Price
3. Критический остаток      (kind=critical_stock)   — нужна срочная поставка
4. Замороженные остатки     (kind=dead_inventory)   — coverage > 180д

Что Александр попросил убрать из Excel:
- low_stock         — дублирует Критический остаток
- repeated_stockout — дублирует Потерянные продажи (это переименованный underestimated)
- sync_error        — теперь шлётся отдельным email из app.notifications:
                       send_sync_error_notification (вызывается из sync.py при ошибке)

Эти три kind остаются в notification_subscriptions для backward compat (если
юзер сам подписан) но в xlsx уже не идут — _build_xlsx их игнорирует.
"""
from __future__ import annotations

import io
import logging
from collections import defaultdict
from datetime import date, datetime, timezone
from typing import Any, Optional

from app.db import fetch_all, get_supabase
from app.jobs.period_window import latest_30d_window as _latest_30d_window, store_metric_30d

logger = logging.getLogger("veloseller.reports")

STORAGE_BUCKET = "report-files"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ─── Форматирование ───────────────────────

def _format_money(value: Any, currency: str = "RUB") -> str:
    if value is None:
        return "—"
    try:
        num = float(value)
    except (TypeError, ValueError):
        return "—"
    sign = "₽" if currency == "RUB" else currency
    return f"{num:,.0f} {sign}".replace(",", " ")


def _bold(cell, color: str = "0F172A") -> None:
    from openpyxl.styles import Font
    cell.font = Font(bold=True, color=color)


def _column_widths(ws, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ─── Метаданные kinds ─────────────────────

SHEET_ROW_LIMIT = 500


# Маппинг kind БД → название листа (тексты Александра 01.06.2026)
KIND_LABELS = {
    "low_stock":          "Низкий остаток",
    "critical_stock":     "Критический остаток",
    "dead_inventory":     "Замороженные остатки",    # переименовано
    "repeated_stockout":  "Частый out-of-stock",
    "underestimated_sku": "Потерянные продажи",       # переименовано
    "sync_error":         "Ошибки синхронизации",
    "weekly_report":      "Сводка по складу",
}

# Какие kinds попадают в Excel.
# sync_error отправляется отдельным email из sync.py при ошибке.
# low_stock и repeated_stockout убраны как дубли по решению Александра.
KINDS_IN_XLSX: frozenset[str] = frozenset({
    "weekly_report",
    "underestimated_sku",
    "critical_stock",
    "dead_inventory",
})

# Порядок листов в Excel (от Александра).
SHEET_ORDER: list[str] = [
    "weekly_report",       # 1. Сводка по складу
    "underestimated_sku",  # 2. Потерянные продажи
    "critical_stock",      # 3. Критический остаток
    "dead_inventory",      # 4. Замороженные остатки
]


def _sheet_name(kind: str) -> str:
    return KIND_LABELS.get(kind, kind)[:31]


# ─── Fetchers ────────────────────────────────

# _latest_30d_window и store_metric_30d импортированы выше из app.jobs.period_window.


def _fetch_sku_rows(sb, seller_id: str, kind: str, params: dict) -> list[dict]:
    """Возвращает строки для листа Excel в формате [{sku, name, ...}, ...]."""
    try:
        # Все kind на основе tvelo_metrics фиксируем на ОДНОМ окне (последний
        # period_end + ~30 дней), иначе SKU троится по числу окон (7/30/90).
        if kind in ("low_stock", "critical_stock", "dead_inventory", "repeated_stockout", "underestimated_sku"):
            period_start, period_end = _latest_30d_window(sb, seller_id)
            if not period_start or not period_end:
                return []
        if kind == "low_stock":
            threshold = int(params.get("coverage_days_threshold", 7))
            res = (
                sb.table("tvelo_metrics")
                .select("coverage_days,current_stock,adjusted_velocity,products!inner(sku,product_name,seller_id,connection_id)")
                .eq("products.seller_id", seller_id)
                .eq("period_start", period_start)
                .eq("period_end", period_end)
                .lte("coverage_days", threshold)
                .gt("current_stock", 0)
                .gt("adjusted_velocity", 0)
                .order("coverage_days")
                .limit(SHEET_ROW_LIMIT)
                .execute()
            )
            return res.data or []

        if kind == "critical_stock":
            # Критический остаток — мало товара, нужна срочная поставка.
            # Колонки Александра: SKU / Название / TVelo / Остаток / Покрытие / Рекомендуемая закупка 30 дней
            threshold = int(params.get("coverage_days_threshold", 3))
            res = (
                sb.table("tvelo_metrics")
                .select("coverage_days,current_stock,current_price,adjusted_velocity,products!inner(sku,product_name,seller_id,connection_id)")
                .eq("products.seller_id", seller_id)
                .eq("period_start", period_start)
                .eq("period_end", period_end)
                .lte("coverage_days", threshold)
                .gt("current_stock", 0)
                .gt("adjusted_velocity", 0)
                .order("coverage_days")
                .limit(SHEET_ROW_LIMIT)
                .execute()
            )
            return res.data or []

        if kind == "dead_inventory":
            # Замороженные остатки — низкая скорость TVelo, деньги стоят.
            # Колонки Александра: SKU / Название / Остаток / TVelo / Покрытие / Заморожено ₽
            threshold = int(params.get("coverage_days_threshold", 180))
            res = (
                sb.table("tvelo_metrics")
                .select("coverage_days,adjusted_velocity,current_stock,current_price,inventory_segment,products!inner(sku,product_name,seller_id,connection_id)")
                .eq("products.seller_id", seller_id)
                .eq("period_start", period_start)
                .eq("period_end", period_end)
                .or_(f"coverage_days.gt.{threshold},inventory_segment.eq.dead_inventory_risk")
                .order("coverage_days", desc=True)
                .limit(SHEET_ROW_LIMIT)
                .execute()
            )
            return res.data or []

        if kind == "repeated_stockout":
            threshold = int(params.get("stockout_days_threshold", 3))
            res = (
                sb.table("tvelo_metrics")
                .select("stockout_days,adjusted_velocity,coverage_days,products!inner(sku,product_name,seller_id,connection_id)")
                .eq("products.seller_id", seller_id)
                .eq("period_start", period_start)
                .eq("period_end", period_end)
                .gte("stockout_days", threshold)
                .order("stockout_days", desc=True)
                .limit(SHEET_ROW_LIMIT)
                .execute()
            )
            return res.data or []

        if kind == "underestimated_sku":
            # Потерянные продажи (бывш. Недооценённый SKU).
            # Колонки Александра: SKU / Название / TVelo / OOS дней / Потеряно ₽
            res = (
                sb.table("tvelo_metrics")
                .select("adjusted_velocity,median_30d_velocity,stockout_days,current_price,products!inner(sku,product_name,seller_id,connection_id),underestimated_sku")
                .eq("products.seller_id", seller_id)
                .eq("period_start", period_start)
                .eq("period_end", period_end)
                .eq("underestimated_sku", True)
                .order("adjusted_velocity", desc=True)
                .limit(SHEET_ROW_LIMIT)
                .execute()
            )
            return res.data or []

        if kind == "sync_error":
            res = (
                sb.table("data_connections")
                .select("name,source,marketplace,last_error,last_sync_at,status")
                .eq("seller_id", seller_id)
                .eq("status", "error")
                .order("last_sync_at", desc=True)
                .limit(SHEET_ROW_LIMIT)
                .execute()
            )
            return res.data or []

        if kind == "weekly_report":
            # Сводка — одна точка. store_metrics пишется по окнам 7/30/90; берём
            # 30-дневное окно последнего периода (иначе .limit(1) брал случайное окно).
            row = store_metric_30d(sb, seller_id, date.today())
            return [row] if row else []

    except Exception:
        logger.exception("_fetch_sku_rows failed kind=%s seller=%s", kind, seller_id)

    return []


# ─── Sheet builders ──────────────────────────────────────

def _row_product(r: dict) -> tuple[str, str]:
    p = r.get("products") or {}
    if isinstance(p, list):
        p = p[0] if p else {}
    return (p.get("sku") or "—", p.get("product_name") or "—")


def _row_warehouse_id(r: dict) -> Optional[str]:
    """connection_id товара из встроенного products — для разбивки отчёта по складам."""
    p = r.get("products") or {}
    if isinstance(p, list):
        p = p[0] if p else {}
    return p.get("connection_id")


def _calc_lost_revenue(r: dict) -> float:
    """Формула Александра: TVelo × OOS дней × Price."""
    return (
        float(r.get("adjusted_velocity") or 0)
        * float(r.get("stockout_days") or 0)
        * float(r.get("current_price") or 0)
    )


def _calc_frozen_money(r: dict) -> float:
    """Замороженные деньги = current_stock × current_price."""
    return float(r.get("current_stock") or 0) * float(r.get("current_price") or 0)


def _build_sheet_weekly_summary(wb, rows: list[dict], currency: str) -> None:
    """Сводка по складу — лист первым.

    Александр 01.06.2026: формат — карточка с числами за неделю, не таблица.
    Берём последнюю запись из store_metrics и оформляем как заглавную страницу
    отчёта: Health Score, потери, заморожено, счётчики SKU.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(_sheet_name("weekly_report"))

    m = rows[0] if rows else {}
    period_end = (m.get("period_end") or "")[:10] if m else "—"

    title_font = Font(bold=True, size=14, color="0F172A")
    label_font = Font(size=10, color="64748B")
    value_font_big = Font(bold=True, size=18, color="0F172A")
    value_font_warn = Font(bold=True, size=18, color="DC2626")
    section_font = Font(bold=True, size=11, color="0F766E")
    section_fill = PatternFill("solid", fgColor="F0FDF4")

    # Заголовок
    ws["A1"] = "Сводка по складу"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Расчёты за неделю · по состоянию на {period_end}"
    ws["A2"].font = label_font
    ws.merge_cells("A2:C2")

    # Health Score крупно
    ws["A4"] = "Health Score"
    ws["A4"].font = section_font
    ws["A4"].fill = section_fill
    ws.merge_cells("A4:C4")

    health = m.get("warehouse_health_score")
    health_val = round(float(health), 0) if health is not None else None
    ws["A5"] = "Здоровье склада"
    ws["A5"].font = label_font
    ws["B5"] = f"{int(health_val)}/100" if health_val is not None else "—"
    ws["B5"].font = value_font_big

    # Деньги
    ws["A7"] = "Деньги"
    ws["A7"].font = section_font
    ws["A7"].fill = section_fill
    ws.merge_cells("A7:C7")

    ws["A8"] = "Потеряно выручки"
    ws["A8"].font = label_font
    ws["B8"] = _format_money(m.get("lost_revenue"), currency)
    ws["B8"].font = value_font_warn

    ws["A9"] = "Заморожено в остатках"
    ws["A9"].font = label_font
    ws["B9"] = _format_money(m.get("store_frozen_inventory_value"), currency)
    ws["B9"].font = value_font_warn

    ws["A10"] = "Стоимость остатков всего"
    ws["A10"].font = label_font
    ws["B10"] = _format_money(m.get("total_inventory_value"), currency)
    ws["B10"].font = value_font_big

    # SKU
    ws["A12"] = "SKU"
    ws["A12"].font = section_font
    ws["A12"].fill = section_fill
    ws.merge_cells("A12:C12")

    ws["A13"] = "Всего SKU"
    ws["A13"].font = label_font
    ws["B13"] = m.get("total_sku_count") or 0
    ws["B13"].font = value_font_big

    ws["A14"] = "В OOS (нет в наличии)"
    ws["A14"].font = label_font
    ws["B14"] = m.get("oos_sku_count") or 0
    ws["B14"].font = value_font_warn

    ws["A15"] = "В замороженных остатках"
    ws["A15"].font = label_font
    ws["B15"] = m.get("dead_inventory_sku_count") or 0
    ws["B15"].font = value_font_warn

    ws["A16"] = "Без активности"
    ws["A16"].font = label_font
    ws["B16"] = m.get("inactive_sku_count") or 0
    ws["B16"].font = value_font_big

    # Концентрация (если есть)
    if m.get("inventory_concentration_50") is not None or m.get("demand_concentration_50") is not None:
        ws["A18"] = "Концентрация"
        ws["A18"].font = section_font
        ws["A18"].fill = section_fill
        ws.merge_cells("A18:C18")

        ws["A19"] = "50% остатков в SKU"
        ws["A19"].font = label_font
        ws["B19"] = m.get("inventory_concentration_50") or "—"
        ws["B19"].font = value_font_big

        ws["A20"] = "50% спроса в SKU"
        ws["A20"].font = label_font
        ws["B20"] = m.get("demand_concentration_50") or "—"
        ws["B20"].font = value_font_big

    _column_widths(ws, {"A": 32, "B": 22, "C": 6})

    for row_num in range(1, 22):
        ws.row_dimensions[row_num].height = 22

    ws.sheet_view.showGridLines = False


def _build_sheet_lost_sales(wb, rows: list[dict], currency: str,
                            sheet_name: Optional[str] = None, warehouse: Optional[str] = None) -> None:
    """Потерянные продажи (kind=underestimated_sku).

    Александр 01.06.2026: SKU / Название / TVelo / OOS дней / Потеряно ₽
    """
    ws = wb.create_sheet(sheet_name or _sheet_name("underestimated_sku"))

    # Подзаголовок-описание (с пометкой склада — отчёт разбит по складам)
    desc = "Товар продаётся быстро, нет в наличии. Каждый день — недополученная выручка."
    if warehouse:
        desc = f"Склад: {warehouse}. " + desc
    ws.append([desc])
    from openpyxl.styles import Font
    ws["A1"].font = Font(italic=True, color="64748B", size=10)
    ws.merge_cells("A1:E1")

    headers = ["SKU", "Название", "TVelo", "OOS дней", f"Потеряно ({currency})"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _bold(ws.cell(row=2, column=col))

    for r in rows:
        sku, name = _row_product(r)
        ws.append([
            sku, name,
            round(float(r.get("adjusted_velocity") or 0), 2),
            int(r.get("stockout_days") or 0),
            round(_calc_lost_revenue(r), 0),
        ])

    _column_widths(ws, {"A": 22, "B": 44, "C": 12, "D": 12, "E": 18})
    ws.freeze_panes = "A3"


def _build_sheet_critical_stock(wb, rows: list[dict], currency: str,
                                sheet_name: Optional[str] = None, warehouse: Optional[str] = None) -> None:
    """Критический остаток.

    Александр 01.06.2026: SKU / Название / TVelo / Остаток / Покрытие / Рекомендуемая закупка 30 дней
    """
    ws = wb.create_sheet(sheet_name or _sheet_name("critical_stock"))

    desc = "Мало товара на складе. Срочно нужна поставка."
    if warehouse:
        desc = f"Склад: {warehouse}. " + desc
    ws.append([desc])
    from openpyxl.styles import Font
    ws["A1"].font = Font(italic=True, color="64748B", size=10)
    ws.merge_cells("A1:F1")

    headers = ["SKU", "Название", "TVelo", "Остаток", "Покрытие (дн)", "Рекомендуемая закупка (30 дн)"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _bold(ws.cell(row=2, column=col))

    for r in rows:
        sku, name = _row_product(r)
        velocity = float(r.get("adjusted_velocity") or 0)
        recommended = round(velocity * 30)
        ws.append([
            sku, name,
            round(velocity, 2),
            int(r.get("current_stock") or 0),
            int(r.get("coverage_days") or 0),
            recommended,
        ])

    _column_widths(ws, {"A": 22, "B": 44, "C": 12, "D": 12, "E": 14, "F": 28})
    ws.freeze_panes = "A3"


def _build_sheet_frozen_stock(wb, rows: list[dict], currency: str,
                              sheet_name: Optional[str] = None, warehouse: Optional[str] = None) -> None:
    """Замороженные остатки (kind=dead_inventory).

    Александр 01.06.2026: SKU / Название / Остаток / TVelo / Покрытие / Заморожено ₽
    """
    ws = wb.create_sheet(sheet_name or _sheet_name("dead_inventory"))

    desc = "Низкая скорость продаж. Деньги заморожены в товаре. Расчёт по средней скорости TVelo за 30 дней."
    if warehouse:
        desc = f"Склад: {warehouse}. " + desc
    ws.append([desc])
    from openpyxl.styles import Font
    ws["A1"].font = Font(italic=True, color="64748B", size=10)
    ws.merge_cells("A1:F1")

    headers = ["SKU", "Название", "Остаток", "TVelo", "Покрытие (дн)", f"Заморожено ({currency})"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        _bold(ws.cell(row=2, column=col))

    for r in rows:
        sku, name = _row_product(r)
        ws.append([
            sku, name,
            int(r.get("current_stock") or 0),
            round(float(r.get("adjusted_velocity") or 0), 2),
            (int(r["coverage_days"]) if r.get("coverage_days") is not None else "∞"),
            round(_calc_frozen_money(r), 0),
        ])

    _column_widths(ws, {"A": 22, "B": 44, "C": 12, "D": 12, "E": 14, "F": 18})
    ws.freeze_panes = "A3"


def _build_sheet_for_kind(wb, kind: str, rows: list[dict], currency: str,
                          sheet_name: Optional[str] = None, warehouse: Optional[str] = None) -> None:
    """Dispatch на правильный билдер по kind."""
    if kind == "weekly_report":
        _build_sheet_weekly_summary(wb, rows, currency)
    elif kind == "underestimated_sku":
        _build_sheet_lost_sales(wb, rows, currency, sheet_name, warehouse)
    elif kind == "critical_stock":
        _build_sheet_critical_stock(wb, rows, currency, sheet_name, warehouse)
    elif kind == "dead_inventory":
        _build_sheet_frozen_stock(wb, rows, currency, sheet_name, warehouse)
    # Остальные kinds в Excel не идут (см. KINDS_IN_XLSX).


# Короткие лейблы листов для разбивки по складам: полное название + имя склада
# не влезают в лимит Excel (31 символ), поэтому для per-склад листов — сокращения.
_KIND_SHORT_LABELS: dict[str, str] = {
    "underestimated_sku": "Потери",
    "critical_stock":     "Критич.",
    "dead_inventory":     "Заморож.",
}


def _unique_sheet_title(wb, base: str) -> str:
    """Имя листа Excel: обрезаем до 31 символа и гарантируем уникальность в книге."""
    title = base[:31]
    if title not in wb.sheetnames:
        return title
    i = 2
    while True:
        suffix = f" {i}"
        candidate = base[:31 - len(suffix)] + suffix
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def _build_xlsx(kind_rows: dict[str, list[dict]], currency: str,
                wh_names: Optional[dict[str, str]] = None) -> bytes:
    """Собирает xlsx из набора kind→rows.

    Склады НЕ смешиваются (решение заказчика 29.06): SKU-листы разбиваются по
    складам — отдельный лист на каждый склад ("<kind> · <Склад>"). Один и тот же
    товар на разных складах больше не выглядит дублем в одной таблице. Сводка
    (weekly_report) остаётся одна — это роллап по магазину, а не список SKU.
    Если у строк нет connection_id (нет инфы о складе) — лист один со старым
    именем (бэк-совместимость).

    Порядок листов — SHEET_ORDER. kinds не из KINDS_IN_XLSX игнорируются.
    weekly_report генерится даже без данных (HEAD-страница «метрик ещё нет»).
    """
    from openpyxl import Workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    names = wh_names or {}
    has_data = False
    for kind in SHEET_ORDER:
        if kind not in KINDS_IN_XLSX or kind not in kind_rows:
            continue
        rows = kind_rows.get(kind) or []

        if kind == "weekly_report":
            _build_sheet_for_kind(wb, kind, rows, currency)
            has_data = True
            continue

        # «Замороженные деньги»: строки с нулём (например WB FBS без цены) — это
        # НЕ замороженные деньги, в этот лист не идут (раньше плодили фантом-дубли).
        if kind == "dead_inventory":
            rows = [r for r in rows if _calc_frozen_money(r) > 0]
        if not rows:
            continue

        # Разбивка по складам — склады не смешиваем.
        groups: dict[Optional[str], list[dict]] = defaultdict(list)
        for r in rows:
            groups[_row_warehouse_id(r)].append(r)

        if list(groups.keys()) == [None]:
            # Нет инфы о складе (edge/тесты) — один лист со старым названием.
            _build_sheet_for_kind(wb, kind, rows, currency)
            has_data = True
            continue

        short = _KIND_SHORT_LABELS.get(kind, _sheet_name(kind))
        for cid in sorted(groups, key=lambda c: (names.get(c or "") or str(c or ""))):
            label = names.get(cid or "") or "Склад"
            title = _unique_sheet_title(wb, f"{short} · {label}")
            _build_sheet_for_kind(wb, kind, groups[cid], currency,
                                  sheet_name=title, warehouse=label)
            has_data = True

    if not has_data:
        ws = wb.create_sheet("Пусто")
        ws.append(["Нет данных для отчётов за этот период."])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Storage upload ───────────────────────────────────

def _upload_xlsx_to_storage(
    sb,
    seller_id: str,
    today_str: str,
    filename: str,
    xlsx_bytes: bytes,
) -> Optional[str]:
    """Заливает XLSX в bucket report-files. Возвращает storage_path или None."""
    path = f"{seller_id}/{today_str}/{filename}"
    try:
        sb.storage.from_(STORAGE_BUCKET).upload(
            path=path,
            file=xlsx_bytes,
            file_options={
                "content-type": XLSX_MIME,
                "upsert": "true",
            },
        )
        return path
    except Exception:
        logger.exception("storage upload failed seller=%s path=%s", seller_id, path)
        return None


# ─── Dispatcher ────────────────────────────────────────

def _today_iso_date() -> str:
    return datetime.now(timezone.utc).date().isoformat()


def _already_sent_today(sb, seller_id: str, channel: str) -> bool:
    try:
        res = (
            sb.table("report_history")
            .select("id")
            .eq("seller_id", seller_id)
            .eq("channel", channel)
            .eq("sent_date", _today_iso_date())
            # Блокируем повтор только на терминальных исходах: 'sent' (доставлено)
            # и 'skipped' (нет данных — ретраить бессмысленно). 'failed' (транзиентный
            # сбой Resend/Telegram) НЕ блокирует — отчёт ретраится в тот же день.
            .in_("status", ["sent", "skipped"])
            .limit(1)
            .execute()
        )
        return bool(res.data)
    except Exception:
        logger.exception("idempotency check failed seller=%s", seller_id)
        return False


def _record_history(
    sb,
    seller_id: str,
    day_of_week: int,
    kinds: list[str],
    channel: str,
    status: str,
    sku_counts: dict[str, int],
    filename: Optional[str],
    file_size: Optional[int],
    storage_path: Optional[str],
    error: Optional[str],
) -> None:
    payload = {
        "seller_id": seller_id,
        "day_of_week": day_of_week,
        "kinds": kinds,
        "channel": channel,
        "status": status,
        "sku_counts": sku_counts,
        "file_name": filename,
        "file_size_bytes": file_size,
        "storage_path": storage_path,
        "error_message": error,
    }
    try:
        sb.table("report_history").insert(payload).execute()
    except Exception:
        # Строка за сегодня уже есть (UNIQUE seller+channel+sent_date+monthly-флаг) —
        # это повторная попытка после 'failed'. Обновляем существующую строку, а не
        # плодим дубль и не ловим повторно unique-violation (иначе отчёт ушёл бы,
        # но статус 'sent' не записался → следующий прогон отправил бы повторно).
        try:
            upd = {k: payload[k] for k in (
                "day_of_week", "kinds", "status", "sku_counts",
                "file_name", "file_size_bytes", "storage_path", "error_message",
            )}
            q = (
                sb.table("report_history").update(upd)
                .eq("seller_id", seller_id)
                .eq("channel", channel)
                .eq("sent_date", _today_iso_date())
            )
            # Сопоставляем тот же monthly-класс, что и в unique-индексе.
            if "monthly_report" in kinds:
                q = q.filter("kinds", "cs", "{monthly_report}")
            else:
                q = q.filter("kinds", "not.cs", "{monthly_report}")
            q.execute()
        except Exception:
            logger.exception("failed to upsert report_history seller=%s", seller_id)


def _warehouse_names(sb, seller_id: str) -> dict[str, str]:
    """connection_id → имя склада (для разбивки отчёта по складам). Ошибки глушим."""
    try:
        res = (
            sb.table("data_connections")
            .select("id,name")
            .eq("seller_id", seller_id)
            .execute()
        )
        return {row["id"]: (row.get("name") or "Склад") for row in (res.data or [])}
    except Exception:
        logger.exception("warehouse names fetch failed seller=%s", seller_id)
        return {}


def dispatch_daily_reports() -> None:
    try:
        sb = get_supabase()
        now_utc = datetime.now(timezone.utc)
        today_dow = now_utc.isoweekday()
        today_dom = now_utc.day

        all_subs = fetch_all(
            sb.table("notification_subscriptions")
            .select("seller_id,kind,channel,enabled,params,frequency")
            .eq("enabled", True)
        )

        groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
        for sub in all_subs:
            # Фильтруем kinds которые больше не идут в Excel
            # (low_stock, repeated_stockout, sync_error).
            # Эти подписки в БД остаются (юзер мог их включить), но мы их
            # тихо пропускаем — Александр попросил убрать дубли.
            if sub["kind"] not in KINDS_IN_XLSX:
                continue

            params = sub.get("params") or {}
            try:
                dow = int(params.get("day_of_week", 1))
            except (ValueError, TypeError):
                dow = 1
            frequency = sub.get("frequency") or "weekly"

            if frequency == "daily":
                pass
            else:
                if dow != today_dow:
                    continue
                if frequency == "monthly" and today_dom > 7:
                    continue

            groups[(sub["seller_id"], sub["channel"])].append(sub)

        if not groups:
            logger.info("dispatch_daily_reports: nothing scheduled for dow=%d dom=%d", today_dow, today_dom)
            return

        sent_email = 0
        sent_telegram = 0
        skipped = 0
        failed = 0

        for (seller_id, channel), subs_list in groups.items():
            if _already_sent_today(sb, seller_id, channel):
                logger.info("skip (already sent today) seller=%s channel=%s", seller_id, channel)
                skipped += 1
                continue

            try:
                seller_res = (
                    sb.table("sellers")
                    .select("id,email,display_name,currency,telegram_chat_id,notify_email,notify_telegram")
                    .eq("id", seller_id)
                    .single()
                    .execute()
                )
                seller = seller_res.data
            except Exception:
                logger.exception("seller fetch failed %s", seller_id)
                failed += 1
                continue

            if not seller:
                skipped += 1
                continue

            if channel == "email" and not seller.get("notify_email", True):
                skipped += 1
                continue
            if channel == "telegram" and not seller.get("notify_telegram", True):
                skipped += 1
                continue

            currency = seller.get("currency") or "RUB"
            kinds = sorted({s["kind"] for s in subs_list})

            kind_rows: dict[str, list[dict]] = {}
            sku_counts: dict[str, int] = {}
            params_by_kind = {s["kind"]: (s.get("params") or {}) for s in subs_list}
            for kind in kinds:
                rows = _fetch_sku_rows(sb, seller_id, kind, params_by_kind[kind])
                kind_rows[kind] = rows
                # weekly_report = HEAD-страница, считаем как 1 запись если есть metrics
                sku_counts[kind] = len(rows)

            # weekly_report сам по себе не "SKU", но если есть подписка только на
            # неё одну — всё равно отправляем (это сводка). Иначе skipped если
            # все sku-листы пусты.
            non_summary_total = sum(
                cnt for kind, cnt in sku_counts.items() if kind != "weekly_report"
            )
            has_summary = sku_counts.get("weekly_report", 0) > 0
            if non_summary_total == 0 and not has_summary:
                logger.info("skip (no data) seller=%s channel=%s kinds=%s",
                            seller_id, channel, kinds)
                _record_history(sb, seller_id, today_dow, kinds, channel,
                                "skipped", sku_counts, None, None, None, "no data")
                skipped += 1
                continue

            try:
                wh_names = _warehouse_names(sb, seller_id)
                xlsx_bytes = _build_xlsx(kind_rows, currency, wh_names)
            except Exception:
                logger.exception("xlsx build failed seller=%s", seller_id)
                _record_history(sb, seller_id, today_dow, kinds, channel,
                                "failed", sku_counts, None, None, None, "xlsx build error")
                failed += 1
                continue

            today_str = date.today().isoformat()
            filename = f"veloseller-otchet-{today_str}.xlsx"

            storage_path = _upload_xlsx_to_storage(sb, seller_id, today_str, filename, xlsx_bytes)

            success = False
            error_msg: Optional[str] = None
            if channel == "email":
                if seller.get("email"):
                    try:
                        from app.notifications import send_report_email
                        success, send_err = send_report_email(
                            to_email=seller["email"],
                            seller_name=seller.get("display_name"),
                            kinds=kinds,
                            sku_counts=sku_counts,
                            xlsx_bytes=xlsx_bytes,
                            filename=filename,
                        )
                        if not success and send_err:
                            error_msg = send_err[:200]
                    except Exception as e:
                        logger.exception("send email failed %s", seller_id)
                        error_msg = f"{type(e).__name__}: {str(e)[:180]}"
                else:
                    error_msg = "no email"
            elif channel == "telegram":
                if seller.get("telegram_chat_id"):
                    try:
                        from app.telegram import send_document, clear_dead_telegram
                        caption = _build_telegram_caption(kinds, sku_counts)
                        success = send_document(
                            chat_id=seller["telegram_chat_id"],
                            file_bytes=xlsx_bytes,
                            filename=filename,
                            caption=caption,
                            on_dead_chat=lambda: clear_dead_telegram(sb, seller_id),
                        )
                    except Exception as e:
                        logger.exception("send telegram failed %s", seller_id)
                        error_msg = f"{type(e).__name__}: {str(e)[:180]}"
                else:
                    error_msg = "no telegram_chat_id"

            if success:
                _record_history(sb, seller_id, today_dow, kinds, channel,
                                "sent", sku_counts, filename, len(xlsx_bytes),
                                storage_path, None)
                if channel == "email":
                    sent_email += 1
                else:
                    sent_telegram += 1
            else:
                _record_history(sb, seller_id, today_dow, kinds, channel,
                                "failed", sku_counts, filename, len(xlsx_bytes),
                                storage_path, error_msg or "send returned False (no details)")
                failed += 1

        logger.info(
            "dispatch_daily_reports done dow=%d dom=%d groups=%d email=%d tg=%d skipped=%d failed=%d",
            today_dow, today_dom, len(groups), sent_email, sent_telegram, skipped, failed,
        )
    except Exception:
        logger.exception("dispatch_daily_reports crashed")


def _build_telegram_caption(kinds: list[str], sku_counts: dict[str, int]) -> str:
    import html
    lines = ["📊 <b>Veloseller — отчёты</b>", ""]
    for kind in kinds:
        label = html.escape(KIND_LABELS.get(kind, kind))
        n = sku_counts.get(kind, 0)
        if kind == "weekly_report":
            lines.append(f"• {label}")
        elif n > 0:
            lines.append(f"• {label}: <b>{n}</b> SKU")
    return "\n".join(lines)
