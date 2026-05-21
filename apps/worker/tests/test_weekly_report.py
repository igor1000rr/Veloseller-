"""Тесты для еженедельного Excel-отчёта."""
from unittest.mock import MagicMock, patch
import io


def _make_sb_with_data(store_metrics_data, top_losses_data, dead_data):
    """Собирает mock Supabase с правильными ответами для трёх запросов."""
    sb = MagicMock()

    def from_table(name):
        m = MagicMock()
        if name == "store_metrics":
            m.select.return_value.eq.return_value.order.return_value.limit.return_value.execute.return_value.data = store_metrics_data
        elif name == "tvelo_metrics":
            # Два разных запроса к tvelo_metrics: top_losses и dead
            sel = m.select.return_value
            # top losses: select.eq.order.limit
            sel.eq.return_value.order.return_value.limit.return_value.execute.return_value.data = top_losses_data
            # dead: select.eq.gt.order.limit
            sel.eq.return_value.gt.return_value.order.return_value.limit.return_value.execute.return_value.data = dead_data
        return m

    sb.table.side_effect = from_table
    return sb


def test_generate_excel_skips_seller_without_store_metrics():
    sb = _make_sb_with_data([], [], [])
    from app.jobs.weekly_report import _generate_excel_for_seller
    result = _generate_excel_for_seller(sb, "seller-1")
    assert result is None


def test_generate_excel_produces_valid_xlsx():
    """Генерация возвращает байты xlsx с 3 листами."""
    store_metrics = [{
        "period_end": "2026-05-21T00:00:00Z",
        "total_sku_count": 100, "oos_sku_count": 5, "low_stock_sku_count": 10,
        "dead_inventory_sku_count": 3, "warehouse_health_score": 78.5,
        "total_inventory_value": 500000, "store_frozen_inventory_value": 50000,
        "lost_revenue": 25000, "inventory_concentration_50": 20, "demand_concentration_50": 15,
    }]
    top_losses = [{
        "adjusted_velocity": 2.5, "coverage_days": 0, "lost_revenue": 12500,
        "demand_pattern": "steady",
        "products": {"sku": "SKU001", "product_name": "Товар 1"},
    }]
    dead = [{
        "coverage_days": 250, "adjusted_velocity": 0.1, "frozen_inventory_value": 30000,
        "products": {"sku": "SKU999", "product_name": "Неликвид"},
    }]
    sb = _make_sb_with_data(store_metrics, top_losses, dead)
    from app.jobs.weekly_report import _generate_excel_for_seller
    result = _generate_excel_for_seller(sb, "seller-1")
    assert result is not None
    assert isinstance(result, bytes)
    assert len(result) > 1000  # реальный xlsx не меньше ~3KB
    # Проверяем что это валидный zip-файл (xlsx — это zip)
    assert result[:2] == b"PK"


def test_generate_excel_with_three_sheets():
    """Проверяем что в xlsx ровно 3 листа."""
    from openpyxl import load_workbook
    store_metrics = [{"period_end": "2026-05-21T00:00:00Z", "total_sku_count": 1, "oos_sku_count": 0,
                      "low_stock_sku_count": 0, "dead_inventory_sku_count": 0,
                      "warehouse_health_score": 100, "total_inventory_value": 0,
                      "store_frozen_inventory_value": 0, "lost_revenue": 0,
                      "inventory_concentration_50": 0, "demand_concentration_50": 0}]
    sb = _make_sb_with_data(store_metrics, [], [])
    from app.jobs.weekly_report import _generate_excel_for_seller
    result = _generate_excel_for_seller(sb, "seller-1")
    assert result is not None
    wb = load_workbook(io.BytesIO(result))
    assert wb.sheetnames == ["Сводка", "Топ потерь", "Неликвид"]
