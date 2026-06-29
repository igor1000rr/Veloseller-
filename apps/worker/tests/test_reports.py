"""Тесты apps/worker/app/jobs/reports.py — диспетчер Excel-отчётов.

Покрытие: _build_xlsx (Excel-генерация), _build_telegram_caption (HTML caption),
dispatch_daily_reports (smoke + idempotency).

Обновлено 01.06.2026 (Veloseller_Отчёт.txt): структура листов поменялась —
low_stock, repeated_stockout, sync_error больше не идут в xlsx (см. KINDS_IN_XLSX).
Активные kinds: weekly_report, underestimated_sku, critical_stock, dead_inventory.
Лейблы: dead_inventory → "Замороженные остатки", underestimated_sku → "Потерянные продажи".
"""
from __future__ import annotations

import io
from unittest.mock import MagicMock, patch

import pytest

from types import SimpleNamespace


# ─── _latest_30d_window (антидубль: окна 7/30/90 на один period_end) ──────────

class _FakeQuery:
    def __init__(self, data): self._data = data
    def select(self, *a, **k): return self
    def eq(self, *a, **k): return self
    def order(self, *a, **k): return self
    def limit(self, *a, **k): return self
    def execute(self): return SimpleNamespace(data=self._data)


class _FakeSB:
    def __init__(self, data): self._data = data
    def table(self, *a, **k): return _FakeQuery(self._data)


def test_latest_30d_window_picks_30day_window():
    """Из окон 7/30/90 дней на последний period_end отчёт берёт 30-дневное —
    иначе SKU дублируется по числу окон."""
    from app.jobs.reports import _latest_30d_window
    data = [
        {"period_start": "2026-06-17", "period_end": "2026-06-23"},  # ~7д
        {"period_start": "2026-05-25", "period_end": "2026-06-23"},  # ~30д
        {"period_start": "2026-03-26", "period_end": "2026-06-23"},  # ~90д
    ]
    ps, pe = _latest_30d_window(_FakeSB(data), "seller-1")
    assert pe == "2026-06-23"
    assert ps == "2026-05-25"


def test_latest_30d_window_no_metrics():
    from app.jobs.reports import _latest_30d_window
    ps, pe = _latest_30d_window(_FakeSB([]), "seller-1")
    assert ps is None and pe is None


# ─── _build_xlsx ──────────────────────────────────────────────────────────────

class TestBuildXlsx:
    def test_returns_valid_zip_bytes(self):
        """XLSX = ZIP. Магическое число PK\\x03\\x04 в начале.

        Используем critical_stock (активный kind) — low_stock больше не идёт в xlsx.
        """
        from app.jobs.reports import _build_xlsx
        result = _build_xlsx({"critical_stock": [
            {"coverage_days": 3, "current_stock": 10, "current_price": 1000, "adjusted_velocity": 2.5,
             "products": {"sku": "SKU-1", "product_name": "Товар 1"}},
        ]}, currency="RUB")
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"
        assert len(result) > 500

    def test_skips_empty_kinds_for_skuish_kinds(self):
        """Пустые списки SKU → лист не создаётся.

        weekly_report — особый случай: HEAD-страница генерится даже с пустыми
        данными (как индикатор что метрик ещё нет). Поэтому в тесте используем
        только SKU-листы — critical_stock и dead_inventory.

        Если ВСЕ пустые → лист 'Пусто'.
        """
        from openpyxl import load_workbook
        from app.jobs.reports import _build_xlsx

        result = _build_xlsx({"critical_stock": [], "dead_inventory": []}, currency="RUB")
        wb = load_workbook(io.BytesIO(result))
        assert wb.sheetnames == ["Пусто"]

    def test_multiple_kinds_become_separate_sheets(self):
        """Несколько kinds → несколько листов в порядке SHEET_ORDER.

        Порядок Александра (01.06.2026):
        weekly_report → underestimated_sku → critical_stock → dead_inventory
        """
        from openpyxl import load_workbook
        from app.jobs.reports import _build_xlsx

        rows_critical = [{
            "coverage_days": 5, "current_stock": 20, "current_price": 500,
            "adjusted_velocity": 1.0,
            "products": {"sku": "A", "product_name": "Товар A"},
        }]
        rows_frozen = [{
            "coverage_days": 200, "adjusted_velocity": 0.1, "current_stock": 30,
            "current_price": 1000,
            "products": {"sku": "B", "product_name": "Товар B"},
        }]

        result = _build_xlsx({
            "dead_inventory": rows_frozen,
            "critical_stock": rows_critical,
        }, currency="RUB")
        wb = load_workbook(io.BytesIO(result))
        # Новые лейблы Александра 01.06.2026:
        # critical_stock → "Критический остаток" (без изменений)
        # dead_inventory → "Замороженные остатки" (бывш. "Неликвид")
        assert "Критический остаток" in wb.sheetnames
        assert "Замороженные остатки" in wb.sheetnames
        # SHEET_ORDER: critical_stock раньше dead_inventory
        assert wb.sheetnames.index("Критический остаток") < wb.sheetnames.index("Замороженные остатки")

    def test_sku_sheets_split_per_warehouse(self):
        """Решение заказчика 29.06: склады НЕ смешиваются. Один и тот же SKU на
        разных складах → отдельные листы, а не одна таблица с «дублями»."""
        from openpyxl import load_workbook
        from app.jobs.reports import _build_xlsx

        rows_frozen = [
            {"coverage_days": 200, "adjusted_velocity": 0.1, "current_stock": 30,
             "current_price": 1000,
             "products": {"sku": "B", "product_name": "Товар B", "connection_id": "c-ozon"}},
            {"coverage_days": 200, "adjusted_velocity": 0.1, "current_stock": 30,
             "current_price": 2000,
             "products": {"sku": "B", "product_name": "Товар B", "connection_id": "c-wb"}},
        ]
        wh = {"c-ozon": "OZON FBS", "c-wb": "WB FBS"}
        result = _build_xlsx({"dead_inventory": rows_frozen}, currency="RUB", wh_names=wh)
        wb = load_workbook(io.BytesIO(result))
        # Два склада → два отдельных листа, без общего "Замороженные остатки".
        assert any("OZON FBS" in s for s in wb.sheetnames)
        assert any("WB FBS" in s for s in wb.sheetnames)
        assert "Замороженные остатки" not in wb.sheetnames

    def test_frozen_sheet_excludes_zero_money_rows(self):
        """В листе замороженных денег строки с нулевой заморозкой (нет цены —
        напр. WB FBS) не показываются: это не «замороженные деньги», а фантом-дубли."""
        from openpyxl import load_workbook
        from app.jobs.reports import _build_xlsx

        rows = [
            {"coverage_days": 200, "adjusted_velocity": 0.1, "current_stock": 30,
             "current_price": 1000,
             "products": {"sku": "REAL", "product_name": "С ценой", "connection_id": "c1"}},
            {"coverage_days": 200, "adjusted_velocity": 0.1, "current_stock": 30,
             "current_price": 0,
             "products": {"sku": "ZERO", "product_name": "Без цены", "connection_id": "c1"}},
        ]
        result = _build_xlsx({"dead_inventory": rows}, currency="RUB", wh_names={"c1": "OZON FBS"})
        wb = load_workbook(io.BytesIO(result))
        sheet = [s for s in wb.sheetnames if "OZON FBS" in s][0]
        ws = wb[sheet]
        skus = [ws.cell(row=r, column=1).value for r in range(3, ws.max_row + 1)]
        assert "REAL" in skus
        assert "ZERO" not in skus

    def test_lost_sales_sheet_has_correct_columns(self):
        """underestimated_sku теперь "Потерянные продажи" с колонками:
        SKU / Название / TVelo / OOS дней / Потеряно ₽
        """
        from openpyxl import load_workbook
        from app.jobs.reports import _build_xlsx

        rows = [{
            "adjusted_velocity": 2.5, "median_30d_velocity": 1.0,
            "stockout_days": 5, "current_price": 1000,
            "products": {"sku": "SKU-A", "product_name": "Товар A"},
            "underestimated_sku": True,
        }]
        result = _build_xlsx({"underestimated_sku": rows}, currency="RUB")
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Потерянные продажи"]
        # Row 1 — описание, row 2 — заголовки
        assert ws["A2"].value == "SKU"
        assert ws["B2"].value == "Название"
        assert ws["C2"].value == "TVelo"
        assert ws["D2"].value == "OOS дней"
        # Потеряно = TVelo × OOS × Price = 2.5 × 5 × 1000 = 12500
        assert ws["A3"].value == "SKU-A"
        assert ws["E3"].value == 12500.0

    def test_deprecated_kinds_skipped(self):
        """Александр 01.06.2026: low_stock, repeated_stockout, sync_error
        больше не идут в Excel.

        Если передать только их — должен получиться лист "Пусто".
        """
        from openpyxl import load_workbook
        from app.jobs.reports import _build_xlsx

        result = _build_xlsx({
            "low_stock": [{"coverage_days": 3, "current_stock": 10,
                           "adjusted_velocity": 2.5,
                           "products": {"sku": "S", "product_name": "T"}}],
            "sync_error": [{"name": "Ozon", "last_error": "x", "last_sync_at": "2026-05-01"}],
            "repeated_stockout": [{"stockout_days": 5, "adjusted_velocity": 1,
                                    "coverage_days": 3,
                                    "products": {"sku": "S2", "product_name": "T2"}}],
        }, currency="RUB")
        wb = load_workbook(io.BytesIO(result))
        # Все три депрекейтнуты — листов нет
        assert wb.sheetnames == ["Пусто"]
        assert "Низкий остаток" not in wb.sheetnames
        assert "Ошибки синхронизации" not in wb.sheetnames
        assert "Частый out-of-stock" not in wb.sheetnames


# ─── _build_telegram_caption ──────────────────────────────────────────────────

class TestTelegramCaption:
    def test_includes_all_kinds_with_counts(self):
        """Лейблы Александра 01.06.2026."""
        from app.jobs.reports import _build_telegram_caption
        result = _build_telegram_caption(
            kinds=["critical_stock", "dead_inventory"],
            sku_counts={"critical_stock": 5, "dead_inventory": 10},
        )
        assert "Критический остаток" in result
        # Бывш. "Неликвид"
        assert "Замороженные остатки" in result
        assert "<b>5</b>" in result
        assert "<b>10</b>" in result

    def test_weekly_report_shown_without_count(self):
        """weekly_report — это HEAD-страница сводки, не SKU. Показываем без счёта."""
        from app.jobs.reports import _build_telegram_caption
        result = _build_telegram_caption(
            kinds=["weekly_report", "critical_stock"],
            sku_counts={"weekly_report": 1, "critical_stock": 3},
        )
        # Сводка показывается без " — N SKU"
        assert "Сводка по складу" in result
        # А SKU-листы — со счётом
        assert "<b>3</b>" in result

    def test_skips_zero_counts(self):
        """Если SKU=0 в kind, его не показываем в caption."""
        from app.jobs.reports import _build_telegram_caption
        result = _build_telegram_caption(
            kinds=["critical_stock", "dead_inventory"],
            sku_counts={"critical_stock": 5, "dead_inventory": 0},
        )
        assert "Критический остаток" in result
        assert "Замороженные остатки" not in result

    def test_caption_under_1024_chars(self):
        """Telegram caption лимит 1024."""
        from app.jobs.reports import _build_telegram_caption
        all_kinds = ["weekly_report", "underestimated_sku", "critical_stock", "dead_inventory"]
        result = _build_telegram_caption(
            kinds=all_kinds,
            sku_counts={k: 999 for k in all_kinds},
        )
        assert len(result) < 1024


# ─── dispatch_daily_reports — smoke ───────────────────────────────────────────

class TestDispatchDailyReports:
    def test_no_subscriptions_does_nothing(self, monkeypatch):
        """Если в БД нет enabled подписок — лог + ранний выход."""
        mock_sb = MagicMock()
        monkeypatch.setattr("app.jobs.reports.get_supabase", lambda: mock_sb)
        monkeypatch.setattr("app.jobs.reports.fetch_all", lambda q: [])

        from app.jobs.reports import dispatch_daily_reports
        dispatch_daily_reports()  # не должно бросать

    def test_filters_subscriptions_by_today_dow(self, monkeypatch):
        """Подписки с params.day_of_week != сегодня — пропускаются.

        Используем critical_stock — активный kind (low_stock теперь deprecated
        и фильтруется до dow-проверки).
        """
        from datetime import datetime, timezone
        import app.jobs.reports as reports_mod

        # Пусть сегодня = понедельник (isoweekday=1)
        fake_today = datetime(2026, 5, 25, 9, 0, 0, tzinfo=timezone.utc)  # Mon
        assert fake_today.isoweekday() == 1

        class FakeDateTime(datetime):
            @classmethod
            def now(cls, tz=None):
                return fake_today

        monkeypatch.setattr(reports_mod, "datetime", FakeDateTime)

        subs = [
            {"seller_id": "s1", "kind": "critical_stock", "channel": "email",
             "enabled": True, "params": {"day_of_week": 1, "coverage_days_threshold": 7}},
            {"seller_id": "s2", "kind": "critical_stock", "channel": "email",
             "enabled": True, "params": {"day_of_week": 3, "coverage_days_threshold": 7}},
        ]
        monkeypatch.setattr(reports_mod, "fetch_all", lambda q: subs)

        groups_seen = {"sellers": set()}

        def fake_fetch_rows(sb, sid, kind, params):
            groups_seen["sellers"].add(sid)
            return []

        monkeypatch.setattr(reports_mod, "_fetch_sku_rows", fake_fetch_rows)
        monkeypatch.setattr(reports_mod, "_already_sent_today", lambda sb, sid, ch: False)
        monkeypatch.setattr(reports_mod, "_record_history",
                            lambda sb, *a, **kw: None)

        mock_sb = MagicMock()
        seller_single = MagicMock()
        seller_single.execute.return_value.data = {
            "id": "s1", "email": "a@b.com", "display_name": "A",
            "currency": "RUB", "telegram_chat_id": None,
            "notify_email": True, "notify_telegram": True,
        }
        mock_sb.table.return_value.select.return_value.eq.return_value.single.return_value = seller_single
        monkeypatch.setattr(reports_mod, "get_supabase", lambda: mock_sb)

        from app.jobs.reports import dispatch_daily_reports
        dispatch_daily_reports()

        # Только s1 (понедельник) попал в обработку, s2 (среда) отфильтрован
        assert groups_seen["sellers"] == {"s1"}

    def test_deprecated_kinds_filtered_at_start(self, monkeypatch):
        """Александр 01.06.2026: low_stock/repeated_stockout/sync_error
        — устаревшие. Должны фильтроваться даже до проверки dow.
        """
        from datetime import datetime, timezone
        import app.jobs.reports as reports_mod

        fake_today = datetime(2026, 5, 25, 9, 0, 0, tzinfo=timezone.utc)

        class FakeDateTime(datetime):
            @classmethod
            def now(cls, tz=None):
                return fake_today

        monkeypatch.setattr(reports_mod, "datetime", FakeDateTime)

        # Все три депрекейтнутые kinds на сегодняшний dow — должны быть отфильтрованы
        subs = [
            {"seller_id": "s1", "kind": "low_stock", "channel": "email",
             "enabled": True, "params": {"day_of_week": 1}},
            {"seller_id": "s2", "kind": "repeated_stockout", "channel": "email",
             "enabled": True, "params": {"day_of_week": 1}},
            {"seller_id": "s3", "kind": "sync_error", "channel": "email",
             "enabled": True, "params": {"day_of_week": 1}},
        ]
        monkeypatch.setattr(reports_mod, "fetch_all", lambda q: subs)

        fetch_calls = {"n": 0}
        monkeypatch.setattr(
            reports_mod, "_fetch_sku_rows",
            lambda sb, sid, kind, params: fetch_calls.update(n=fetch_calls["n"]+1) or [],
        )
        monkeypatch.setattr(reports_mod, "get_supabase", lambda: MagicMock())

        from app.jobs.reports import dispatch_daily_reports
        dispatch_daily_reports()

        # Никаких fetch'ей — все три kinds отфильтрованы до начала обработки
        assert fetch_calls["n"] == 0

    def test_skipped_when_idempotency_check_hits(self, monkeypatch):
        """Если уже отправляли сегодня — fetcher не вызывается."""
        from datetime import datetime, timezone
        import app.jobs.reports as reports_mod

        fake_today = datetime(2026, 5, 25, 9, 0, 0, tzinfo=timezone.utc)

        class FakeDateTime(datetime):
            @classmethod
            def now(cls, tz=None):
                return fake_today

        monkeypatch.setattr(reports_mod, "datetime", FakeDateTime)

        subs = [{
            "seller_id": "s1", "kind": "critical_stock", "channel": "email",
            "enabled": True, "params": {"day_of_week": 1, "coverage_days_threshold": 7},
        }]
        monkeypatch.setattr(reports_mod, "fetch_all", lambda q: subs)
        monkeypatch.setattr(reports_mod, "_already_sent_today", lambda sb, sid, ch: True)

        fetch_called = {"n": 0}
        monkeypatch.setattr(
            reports_mod, "_fetch_sku_rows",
            lambda sb, sid, kind, params: fetch_called.update(n=fetch_called["n"]+1) or [],
        )
        monkeypatch.setattr(reports_mod, "get_supabase", lambda: MagicMock())

        from app.jobs.reports import dispatch_daily_reports
        dispatch_daily_reports()

        # Если idempotency сработала, fetcher вообще не должен вызываться.
        assert fetch_called["n"] == 0
